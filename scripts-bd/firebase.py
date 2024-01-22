import firebase_admin
from firebase_admin import credentials, firestore
import openpyxl

cred = credentials.Certificate("desafiotecnicogreat.json")
firebase_admin.initialize_app(cred)
db = firestore.client()

def read_excel_file_line_by_line(file_path, sheet_name):

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        row_data = [cell.value for cell in row]
        yield row_data

def write_data_to_firestore(estado,municipio,escola,data):
    collection_ref = db.collection("estados").document(estado).collection(municipio).document("escolas").collection(escola).document("dados")
    collection_ref.set(data)

for line in read_excel_file_line_by_line("sudeste.xlsx", "INSE_ESC_2021"):
    estado = line[0]
    municipio = line[1]
    escola = line[2]

    data = {
        "TP_TIPO_REDE": line[3],
        "TP_LOCALIZACAO": line[4],
        "TP_CAPITAL": line[5],
        "QTD_ALUNOS_INSE": line[6],
        "MEDIA_INSE": line[7],
        "INSE_CLASSIFICACAO": line[8],
        "PC_NIVEL_1": line[9],
        "PC_NIVEL_2": line[10],
        "PC_NIVEL_3": line[11],
        "PC_NIVEL_4": line[12],
        "PC_NIVEL_5": line[13],
        "PC_NIVEL_6": line[14],
        "PC_NIVEL_7": line[15],
        "PC_NIVEL_8": line[16]
    }

    write_data_to_firestore(estado,municipio,escola,data)